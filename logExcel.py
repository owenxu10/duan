# -*- coding: utf-8 -*-
import os
import os.path
import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill
import json
from elasticsearch import Elasticsearch
from util.util import handleQuestion, afterSearch


def style_range(ws):
    border = Border(left=Side(style='thin', color='FF99C4E6'),
                    right=Side(style='thin', color='FF99C4E6'),
                    top=Side(style='thin', color='FF99C4E6'),
                    bottom=Side(style='thin', color='FF99C4E6'))

    fill = PatternFill("solid", fgColor="f9fbfd")
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(column=col, row=row).border = border
            ws.cell(column=col, row=row).fill = fill


def logFiles(rootdir):
    files = []
    dict = {}
    for parent, dirnames, filenames in os.walk(rootdir):
        for filename in filenames:
            num = int(re.match(r'\.\/log_server\/history\.log\.(\d*)', os.path.join(parent, filename)).group(1))
            dict[os.path.join(parent, filename)] = num
    dicts = sorted(dict.items(), key=lambda d: d[1])
    for dict in dicts:
        files.append(dict[0])
    return files


def getSearchResult(answer, questions):
    es = Elasticsearch(hosts='127.0.0.1')
    searchResult = ''

    if len(answer) != 0:
        query = {
            "query": {
                "terms": {
                    "_id": [answer]
                }
            }
        }
        res = es.search(index="qa_demo", body=json.dumps(query, ensure_ascii=False))
        if len(res['hits']['hits']) > 0:
            searchResult = res['hits']['hits'][0]['_source']['answer'] + '\n\n'

    if len(questions) != 0:
        question = questions.split('|')
        query = {
            "query": {
                "terms": {
                    "_id": question
                }
            }
        }

        res = es.search(index="qa_demo", body=json.dumps(query, ensure_ascii=False))
        hits = res['hits']['hits']
        for i in range(0, len(hits)):
            searchResult = searchResult + str(i + 1) + '.' + hits[i]['_source']['question'] + '\n'

    return searchResult + answer + questions


def getRealTimeResult(question):
    group = re.search("(?:\d*)](.*)", question)
    if group is not None:
        question = group.group(1)
    es = Elasticsearch(hosts='127.0.0.1')
    bar = 100
    content = ''
    answer = ''
    # handleQuestion
    (question_sequence, question, period, isSub) = handleQuestion(question)
    if isSub:
        bar = 90
    query_data = {
        "query": {
            "bool": {
                "should": [
                    {
                        "multi_match": {
                            "query": question_sequence,
                            "fields": [
                                "question^2",
                                "answer^1"
                            ],
                            "boost": 1
                        }
                    },
                    {
                        "match_phrase": {
                            "question": {
                                "query": question_sequence,
                                "boost": 3
                            }
                        }
                    },
                    {
                        "match_phrase": {
                            "answer": {
                                "query": question_sequence,
                                "slop": 0,
                                "boost": 2
                            }
                        }
                    }
                ]
            }
        },
        "highlight": {
            "pre_tags": [
                ""
            ],
            "post_tags": [
                ""
            ],
            "fields": {
                "question": {},
                "answer": {}
            }
        }
    }
    res = es.search(index="qa_demo", body=json.dumps(query_data, ensure_ascii=False))

    filtered_hits, maxScore, took = afterSearch(res, period)

    if len(filtered_hits) >= 6:
        if maxScore > bar:
            print(filtered_hits[0]['_source'])
            if filtered_hits[0]['_source']['href'] != None:
                answer = filtered_hits[0]['_source']['answer'] + '\n' \
                         + "【知识来源:段涛大夫" + filtered_hits[0]['_source']['title'] + "】"
            content = answer + "\n\n" \
                      + '1.' + filtered_hits[1]['_source']['question'].strip() + '\n' \
                      + '2.' + filtered_hits[2]['_source']['question'].strip() + '\n' \
                      + '3.' + filtered_hits[3]['_source']['question'].strip() + '\n' \
                      + '4.' + filtered_hits[4]['_source']['question'].strip() + "\n" \
                      + '5.' + filtered_hits[5]['_source']['question'].strip() + "\n"

        else:
            content = '1.' + filtered_hits[0]['_source']['question'].strip() + '\n' \
                      + '2.' + filtered_hits[1]['_source']['question'].strip() + '\n' \
                      + '3.' + filtered_hits[2]['_source']['question'].strip() + '\n' \
                      + '4.' + filtered_hits[3]['_source']['question'].strip() + "\n" \
                      + '5.' + filtered_hits[4]['_source']['question'].strip() + "\n"

    elif len(filtered_hits) == 1:

        if maxScore > bar:
            if filtered_hits[0]['_source']['href'] != None:
                content = filtered_hits[0]['_source']['answer'] + '\n' \
                          + "【知识来源:段涛大夫" + filtered_hits[0]['_source']['title'] + "】"
        else:

            index = 1
            for filtered_hit in filtered_hits:
                content = content + str(index) + '.' + filtered_hit['_source']['question'].strip() + "\n"
                index += 1
    elif len(filtered_hits) > 0:
        if maxScore > bar:
            print(filtered_hits[0]['_source'])
            if filtered_hits[0]['_source']['href'] != None:
                answer = filtered_hits[0]['_source']['answer'] + '\n' \
                         + "【知识来源:段涛大夫" + filtered_hits[0]['_source']['title'] + "】"

            content = answer + "\n\n"
            index = 1
            for index in range(1, len(filtered_hits)):
                content = content + str(index) + '.' + filtered_hits[index]['_source']['question'].strip() + "\n"
                index += 1

        else:
            index = 1
            for filtered_hit in filtered_hits:
                content = content + str(index) + '.' + filtered_hit['_source']['question'].strip() + "\n"
                index += 1

    return content


def readLogs(files):
    currentDate = 0
    answer = ''
    c_count = 0
    c_suc = 0
    c_fail = 0
    l_count = 0
    l_suc = 0
    l_fail = 0
    ll_suc = []
    ll_suc_answer = []
    ll_suc_questions = []
    ll_fail = []

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    for file in files:
        with open(file, 'r', encoding="utf-8") as f:
            print(file)
            for line in f:
                print("line")
                date = line[0:10]
                re_result = re.search("2017-(.*?)-(.*?) ", line)
                if re_result is not None:
                    log_month = int(re_result.group(1))
                    log_day = int(re_result.group(2))
                score_group = re.search("SCORE\[(.*?)\]", line)
                if score_group is not None:
                    score = score_group.group(1)
                search_group = re.search("SEARCH\[(.*?)\]", line)
                if search_group is not None:
                    search = '[' + score + ']' + search_group.group(1)
                answer_group = re.search("ANSWER\[(.*?)\]", line)
                if answer_group is not None:
                    answer = answer_group.group(1)
                questions_group = re.search("RESULT\[(.*?)\]", line)
                questions = ''
                if questions_group is not None:
                    questions = questions_group.group(1)

                if date != currentDate and currentDate != 0:
                    # Write
                    print("currentDate", currentDate)
                    ws = wb.create_sheet(title=currentDate)

                    ws['A1'] = "当日搜索次数:" + str(l_count)
                    ws['A2'] = "成功匹配:" + str(l_suc)
                    ws['C2'] = "未成功匹配:" + str(l_fail)
                    ws['B2'] = "匹配率:" + str(l_suc / l_count * 100) + "%"

                    ws['A3'] = "成功匹配问题"
                    ws['B3'] = "匹配结果"
                    ws['C3'] = "未成功匹配问题"

                    for i in range(0, len(ll_suc)):
                        ws.cell(column=1, row=i + 4).value = ll_suc[i]

                        result = getRealTimeResult(ll_suc[i])
                        # getSearchResult(ll_suc_answer[i], ll_suc_questions[i])
                        ws.cell(column=2, row=i + 4).value = result
                    for i in range(0, len(ll_fail)):
                        ws.cell(column=3, row=i + 4).value = ll_fail[i]

                    style_range(ws)
                    # Clear
                    l_count = l_suc = l_fail = 0
                    ll_suc.clear()
                    ll_fail.clear()
                    ll_suc = []
                    ll_suc_answer = []
                    ll_suc_questions = []
                    ll_fail = []

                c_count += 1
                l_count += 1
                currentDate = date
                if answer == 'None' and len(questions) == 0:
                    print("======")
                    c_fail += 1
                    l_fail += 1
                    ll_fail.append(search)
                else:
                    c_suc += 1
                    l_suc += 1
                    ll_suc.append(search)
                    ll_suc_answer.append(answer)
                    ll_suc_questions.append(questions)

    ws = wb.create_sheet(title=currentDate)

    ws['A1'] = "当日搜索次数:" + str(l_count)
    ws['A2'] = "成功匹配:" + str(l_suc)
    ws['C2'] = "未成功匹配:" + str(l_fail)
    ws['B2'] = "匹配率:" + str(l_suc / l_count * 100) + "%"

    ws['A3'] = "成功匹配问题"
    ws['B3'] = "匹配结果"
    ws['C3'] = "未成功匹配问题"

    for i in range(0, len(ll_suc)):
        ws.cell(column=1, row=i + 4).value = ll_suc[i]

        result = getSearchResult(ll_suc_answer[i], ll_suc_questions[i])
        ws.cell(column=2, row=i + 4).value = result
    for i in range(0, len(ll_fail)):
        ws.cell(column=3, row=i + 4).value = ll_fail[i]

    style_range(ws)

    overview = wb["Overview"]
    overview['A1'] = "所有搜索次数:" + str(c_count)
    overview['A2'] = "成功匹配:" + str(c_suc)
    overview['A3'] = "未成功匹配:" + str(c_fail)
    overview['A4'] = "匹配率:" + str(c_suc / c_count * 100) + "%"

    dest_filename = 'log.xlsx'
    style_range(overview)
    wb.save(filename=dest_filename)


if __name__ == "__main__":
    wb = Workbook()

    rootdir = './log_server'  # log direcory
    files= logFiles(rootdir)
    print(files[::-1])
    readLogs(files)
