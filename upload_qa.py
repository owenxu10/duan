# -*-coding:utf-8 -*-

from elasticsearch import Elasticsearch
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import urllib.request
import time

es = Elasticsearch()
questions = [];
answers = [];
wb = load_workbook('dt.xlsx')
sheet_names = wb.get_sheet_names()
ws = wb[sheet_names[0]]


def delete_es():
    es.indices.delete(index='qa_demo')


def create_mapping():
    es.indices.create('qa_demo', body={
        'mappings': {
            'qa': {
                'properties': {
                    'question': {
                        'type': 'text',
                        'analyzer': 'ik_syno'
                    },
                    'answer': {
                        'type': 'text',
                        'analyzer': 'ik_syno'
                    },
                    'title': {
                        'type': 'text'
                    },
                    'href': {
                        'type': 'text'
                    },
                    'period': {
                        'type': 'text'
                    }
                }
            }
        },
        "settings": {
            "number_of_shards": 1,
            "number_of_replicas": 1,
            "analysis": {
                "filter": {
                    "my_synonym_filter": {
                        "type": "synonym",
                        "synonyms_path": "analysis/synonym.txt"
                    }
                },
                "analyzer": {
                    "ik_syno": {
                        "type": "custom",
                        "tokenizer": "ik_max_word",
                        "filter": "my_synonym_filter"
                    },
                    "ik_syno_smart": {
                        "type": "custom",
                        "tokenizer": "ik_smart",
                        "filter": "my_synonym_filter"
                    }
                }
            }
        }
    })


def upload(question, answer, title, href, period):
    print(locals())
    res = es.index(index='qa_demo', doc_type='qa', body={
        'question': question,
        'answer': answer,
        'title': title,
        'href': href,
        'period': period
    })
    return res


def checkSame(question, answer):
    if (question in questions) and (answer in answers):
        return False
    else:
        questions.append(question)
        answers.append(answer)
        return True


def getTitle(source):
    response = urllib.request.urlopen(source)
    time.sleep(1)
    soup = BeautifulSoup(response.read(), 'lxml')
    title = u"《" + soup.select('h2.rich_media_title')[0].text.strip() + u"》"
    return title


def read_xls_1(ws):
    for i in range(7, 758):

        question = ws.cell(row=i, column=10).value  # start question
        answer = ws.cell(row=i, column=11).value  # start answer
        source = ws.cell(row=i, column=12).value  # start hyperlink
        period = processPeriod(ws.cell(row=i, column=1).value)

        if ws.cell(row=i, column=12).hyperlink == None:
            soup = BeautifulSoup(source, "lxml")

            if soup.a == None:
                href = None
                title = source[source.index(u'《'): source.index(u'》') + 1]
                # print title
            else:
                href = soup.a['href']
                origin_title = soup.a.text
                title = origin_title[origin_title.index(u'《'): origin_title.index(u'》') + 1]
        else:
            href = ws.cell(row=i, column=12).hyperlink.display
            origin_title = ws.cell(row=i, column=12).value
            title = origin_title[origin_title.index(u'《'): origin_title.index(u'》') + 1]

        # True- no same, False -same
        if checkSame(question, answer):
            upload(question, answer, title, href, period)


# new data from 758
def read_xls_2(ws):
    for i in range(758, ws.max_row + 1):
        question = ws.cell(row=i, column=10).value  # start question
        answer = ws.cell(row=i, column=11).value  # start answer
        source = ws.cell(row=i, column=12).value  # start hyperlink
        period = processPeriod(ws.cell(row=i, column=1).value)

        title = getTitle(source)

        # True- no same, False -same
        if checkSame(question, answer):
            upload(question, answer, title, source, period)


def processPeriod(period_origin):
    period = ""

    # 无
    if period_origin == None or "避孕" in period_origin or len(period_origin) == 0:
        period = "无"
    # 备孕
    # 备孕期
    # 孕前
    elif "备孕期" in period_origin or "孕前" in period_origin:
        period = "备孕"
    # 孕期
    # 孕早期 / 孕中期 / 孕晚期
    elif "孕早期" in period_origin or "孕中期" in period_origin  or "孕晚期" in period_origin:
        period = "孕期"
    # 产后
    # 临产
    # 新生儿
    # 流产
    elif "产后" in period_origin or "临产" in period_origin \
            or "新生儿" in period_origin or "流产" in period_origin:
        period = period_origin

    print("period_origin:", period_origin)
    print("period:", period)
    return period


if __name__ == '__main__':
    delete_es()
    create_mapping()
    read_xls_1(ws)
    read_xls_2(ws)
