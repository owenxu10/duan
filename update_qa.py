# -*-coding:utf-8 -*-

from elasticsearch import Elasticsearch
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import urllib.request
import time, json

es = Elasticsearch()
questions = [];
answers = [];
wb = load_workbook('./excels/dt.xlsx')
sheet_names = wb.get_sheet_names()
ws = wb[sheet_names[0]]


def delete_es():
    es.indices.delete(index='qa_demo')


def create_mapping():
    es.indices.create('qa_demo', body={
        'mappings': {
            'qa': {
                'properties': {
                    'question': {
                        'type': 'text',
                        'analyzer': 'ik_syno'
                    },
                    'answer': {
                        'type': 'text',
                        'analyzer': 'ik_syno'
                    },
                    'title': {
                        'type': 'text'
                    },
                    'href': {
                        'type': 'text'
                    },
                    'period': {
                        'type': 'text'
                    }
                }
            }
        },
        "settings": {
            "number_of_shards": 1,
            "number_of_replicas": 1,
            "analysis": {
                "filter": {
                    "my_synonym_filter": {
                        "type": "synonym",
                        "synonyms_path": "analysis/synonym.txt"
                    }
                },
                "analyzer": {
                    "ik_syno": {
                        "type": "custom",
                        "tokenizer": "ik_max_word",
                        "filter": "my_synonym_filter"
                    },
                    "ik_syno_smart": {
                        "type": "custom",
                        "tokenizer": "ik_smart",
                        "filter": "my_synonym_filter"
                    }
                }
            }
        }
    })


def upload(question, answer, title, href, period):
    print(locals())
    print(question.strip())
    if title != None:
        title = title.strip()
    res = es.index(index='qa_demo', doc_type='qa', body={
        'question': question.strip(),
        'answer': answer.strip(),
        'title': title,
        'href': href,
        'period': period
    })
    return res


def checkSame(question, answer):
    noSame = True
    query_all = {
        "query": {
            "bool": {
                "should": [
                    {
                        "match_phrase": {
                            "question": {
                                "query": question
                            }
                        }
                    },
                    {
                        "match_phrase": {
                            "answer": {
                                "query": answer
                            }
                        }
                    }
                ]
            }
        }
    }

    res = es.search(index="qa_demo", body=json.dumps(query_all, ensure_ascii=False))
    if len(res['hits']['hits']) > 0:
        noSame = False
    return noSame


def getTitle(source):
    response = urllib.request.urlopen(source)
    time.sleep(1)
    soup = BeautifulSoup(response.read(), 'lxml')
    try:
        title = u"《" + soup.select('h2.rich_media_title')[0].text.strip() + u"》"
    except:
        title = None
    return title


def read_xls(ws):
    index = 0
    for i in range(841, ws.max_row + 1):
        question = ws.cell(row=i, column=10).value  # start question
        answer = ws.cell(row=i, column=11).value  # start answer
        source = ws.cell(row=i, column=12).value  # start hyperlink
        period = processPeriod(ws.cell(row=i, column=1).value)

        if question == None:
            break;

        if source != None:
            title = getTitle(source)
        print(question)
        # True- no same, False -same
        if checkSame(question, answer):
            index += 1
            print(index)
            upload(question, answer, title, source, period)


def processPeriod(period_origin):
    period = ""

    # 无
    if period_origin == None or "避孕" in period_origin or len(period_origin) == 0:
        period = "无"
    # 备孕
    # 备孕期
    # 孕前
    elif "备孕期" in period_origin or "孕前" in period_origin:
        period = "备孕"
    # 孕期
    # 孕早期 / 孕中期 / 孕晚期
    elif "孕早期" in period_origin or "孕中期" in period_origin or "孕晚期" in period_origin:
        period = "孕期"
    # 产后
    # 临产
    # 新生儿
    # 流产
    elif "产后" in period_origin or "临产" in period_origin \
            or "新生儿" in period_origin or "流产" in period_origin:
        period = period_origin

    return period


if __name__ == '__main__':
    read_xls(ws)
